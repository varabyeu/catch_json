from openpyxl import load_workbook
from pydantic import ValidationError

from .models import ProductData


def get_articles(file):
    """Func returns lists of a correct and incorrect articles

    If an article data type is not an integer, the article moves to a list_error.
    If an article data type is integer, the article moves to a list_articles
    """
    workbook = load_workbook(filename=file).active
    list_articles = []
    list_error_articles = []
    row = 1

    while True:
        cell = workbook.cell(row=row, column=1)
        if cell.value:
            if isinstance(cell.value, int):
                list_articles.append(cell.value)
            else:
                list_error_articles.append(cell.value)
            row += 1
        else:
            break
    return tuple(list_articles), tuple(list_error_articles)


def get_valid_data(response):
    """Func to valid data

    Validate data by pydantic's ProdductData
    """
    try:
        valid_data = ProductData.parse_obj(response)
        return {
            'article': valid_data['nm_id'],
            'brand': valid_data['brand_name'],
            'title': valid_data['imt_name']
        }
    except ValidationError as ve:
        print(f'Validation error during processing {response}')
